from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
from pathlib import Path
from fpdf import FPDF
from PIL import Image
from openpyxl import load_workbook
from flask import current_app
from win32com.client import constants

import pythoncom
import win32com.client as win32
import os, requests, requests, shutil, zipfile, time, pythoncom

# ⬇️ PyInstaller가 같이 묶도록 강제
try:
    import win32timezone  # noqa: F401
except Exception:
    pass

# 새 엑셀파일로 만드는 함수
def make_excel_auto(data_list, filename):
    if not data_list:
        raise ValueError("❌ 빈 리스트입니다. 저장할 데이터가 없습니다.")

    # 새 워크북 생성
    wb = Workbook()
    ws = wb.active

    # 첫 row의 key들을 헤더로 자동 추출
    keys = list(data_list[0].keys())
    ws.append(keys)

    # 각 row의 value 삽입
    for item in data_list:
        row = [item.get(k, "") for k in keys]
        ws.append(row)

    # 파일 이름 및 경로 생성
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{now}_{filename}.xlsx"
    filepath = os.path.abspath(os.path.join("app", "result", filename))

    # 저장
    wb.save(filepath)
    print(f"✅ 자동 엑셀 저장 완료: {filepath}")
    return filepath, filename

# png, jpg 파일을 1개의 pdf 파일로 만드는 함수.
def make_image_pdf(url_list, filename, output_dir="app/result"):
    import os
    import requests
    from fpdf import FPDF
    from PIL import Image

    if not url_list:
        raise ValueError("❌ url_list가 비어있습니다. PDF를 생성할 수 없습니다.")

    pdf = FPDF(unit="mm")  # 단위 설정
    os.makedirs(output_dir, exist_ok=True)

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
    }

    for idx, img_url in enumerate(url_list):
        try:
            # 이미지 다운로드 및 임시 저장
            ext = img_url.split('.')[-1].split("?")[0]
            img_path = f"temp_img_{idx+1:02d}.{ext}"

            response = requests.get(img_url, headers=headers)
            if response.status_code != 200:
                print(f"❌ 이미지 {idx+1} 다운로드 실패: {response.status_code}")
                continue

            with open(img_path, "wb") as f:
                f.write(response.content)

            # 이미지 열기 및 크기 측정
            img = Image.open(img_path).convert("RGB")
            width_px, height_px = img.size
            width_mm = width_px * 0.264583
            height_mm = height_px * 0.264583

            # 기본 A4 페이지에 가운데 맞추기
            pdf.add_page()
            x_offset = max(0, (210 - width_mm) / 2)
            y_offset = max(0, (297 - height_mm) / 2)
            pdf.image(img_path, x=x_offset, y=y_offset, w=width_mm, h=height_mm)

            os.remove(img_path)

        except Exception as e:
            print(f"❌ 이미지 {idx+1} 처리 실패: {e}")
            continue

    # 최종 PDF 저장
    pdf_path = os.path.abspath(os.path.join(output_dir, f"{filename}.pdf"))
    pdf.output(pdf_path)
    print(f"✅ PDF 생성 완료: {pdf_path}")
    return pdf_path

# 개별 pdf 만드는 함수.
def make_multiple_pdfs_and_zip(url_list, name_list, base_filename, output_dir="app/result"):
    """
    이미지 URL 리스트 + 이름 리스트를 받아 각각 PDF로 저장하고 ZIP 압축

    Parameters:
        url_list (list): 이미지 URL 리스트
        name_list (list): 각 이미지에 대응되는 이름 리스트
        base_filename (str): 압축파일 이름
        output_dir (str): 저장 경로

    Returns:
        (str, str): (압축파일 경로, 압축파일 이름)
    """
    if not url_list or not name_list or len(url_list) != len(name_list):
        raise ValueError("❌ url_list와 name_list는 길이가 같아야 합니다.")

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d")
    folder_name = f"{base_filename}_{timestamp}"
    folder_path = os.path.join(output_dir, folder_name)
    os.makedirs(folder_path, exist_ok=True)

    headers = {"User-Agent": "Mozilla/5.0"}

    for idx, (url, name) in enumerate(zip(url_list, name_list)):
        try:
            ext = url.split('.')[-1].split("?")[0]
            img_path = os.path.join(folder_path, f"temp_{idx+1:02d}.{ext}")

            response = requests.get(url, headers=headers)
            if response.status_code != 200:
                print(f"❌ 이미지 {idx+1} 다운로드 실패: {response.status_code}")
                continue

            with open(img_path, "wb") as f:
                f.write(response.content)

            img = Image.open(img_path).convert("RGB")
            width_px, height_px = img.size
            width_mm = width_px * 0.264583
            height_mm = height_px * 0.264583

            pdf = FPDF(unit="mm", format=(width_mm, height_mm))
            pdf.add_page()
            pdf.image(img_path, x=0, y=0, w=width_mm, h=height_mm)

            safe_name = "".join(c for c in name if c.isalnum() or c in (' ', '_', '-')).strip()
            pdf_filename = f"{safe_name}.pdf"
            pdf_path = os.path.join(folder_path, pdf_filename)
            pdf.output(pdf_path)

            os.remove(img_path)

        except Exception as e:
            print(f"❌ 이미지 {idx+1} 처리 실패: {e}")
            continue

    zip_filename = f"{folder_name}.zip"
    zip_path = os.path.abspath(os.path.join(output_dir, zip_filename))

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
        for file in os.listdir(folder_path):
            zipf.write(os.path.join(folder_path, file), arcname=file)

    print(f"✅ 압축 완료: {zip_path}")
    return zip_path, zip_filename

# ================================ '엑셀에 데이터를 덧붙이는 함수' 모음 시작 ================================
def _parse_date(v):
    if v is None or v == "": return None
    if isinstance(v, (datetime, date)): return datetime(v.year, v.month, v.day)
    s = str(v).strip().replace(" ", "")
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try: return datetime.strptime(s, fmt)
        except: pass
    if len(s) == 8 and s.isdigit():
        try: return datetime.strptime(s, "%Y%m%d")
        except: pass
    try:
        iv = float(s)
        base = datetime(1899, 12, 30)
        if 1 <= iv < 600000: return base + timedelta(days=iv)
    except: pass
    return None

def _norm_text(x):
    if x is None: return ""
    s = str(x).replace("\r", "\n")
    s = "\n".join(part.strip() for part in s.split("\n"))
    return " ".join(s.split())  # 한 줄로

# 기존 엑셀파일에 덧붙이기
def append_unique_to_excel(
    data_list,
    filename,
    filepath,
    sheetname="검색",
    col_mapping=None,    # {"q_title":"질문제목", ...}  <-- 열 순서도 이 순서로 고정
    key_fields=None,     # ["q_shopping_mall","q_title","qa_date"] 권장
    sort_by=None,        # "qa_date" 등
    sort_desc=True,
):
    if not col_mapping or not key_fields:
        raise ValueError("col_mapping / key_fields는 필수입니다.")

    path = Path(filepath); path.parent.mkdir(parents=True, exist_ok=True)

    # 파일 없으면 생성
    if not path.exists():
        wb = Workbook(); ws = wb.active; ws.title = sheetname
        ws.append([col_mapping[k] for k in col_mapping.keys()])
        wb.save(path)

    wb = load_workbook(path)
    ws = wb[sheetname] if sheetname in wb.sheetnames else wb.active

    desired_header = [col_mapping[k] for k in col_mapping.keys()]
    header = [c.value for c in ws[1]]
    if header != desired_header:
        ws.delete_rows(1, ws.max_row)
        ws.append(desired_header)

    # 기존 데이터 로드
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r): continue
        d = {}
        for i, k in enumerate(col_mapping.keys()):
            d[k] = r[i] if i < len(r) else None
        rows.append(d)

    def ktuple(d):
        return tuple(_norm_text(d.get(k, "")) for k in key_fields)

    seen = {ktuple(d) for d in rows}

    # 신규 데이터 정규화 후 병합
    for it in data_list:
        rec = {}
        for k in col_mapping.keys():
            v = it.get(k, "")
            if k == sort_by:
                dv = _parse_date(v)
                rec[k] = dv or _norm_text(v)
            else:
                rec[k] = _norm_text(v)
        if ktuple(rec) in seen: continue
        rows.append(rec); seen.add(ktuple(rec))

    # 정렬
    if sort_by:
        def keyfunc(d):
            v = d.get(sort_by)
            if isinstance(v, (datetime, date)): return v
            pv = _parse_date(v); return pv or str(v)
        rows.sort(key=keyfunc, reverse=sort_desc)

    # 시트 다시 쓰기(헤더 제외 전체 갈아끼움)
    ws.delete_rows(2, ws.max_row)
    for d in rows:
        row = []
        for k in col_mapping.keys():
            val = d.get(k, "")
            if k == sort_by:
                dv = _parse_date(val) if not isinstance(val, (datetime, date)) else val
                if dv: 
                    row.append(dv);  # 날짜는 진짜 datetime으로 기록
                else:
                    row.append(val)
            else:
                row.append(val)
        ws.append(row)

    # 서식: 줄바꿈 off, 상단 고정, 폭 조정, 날짜 포맷
    ws.freeze_panes = "A2"
    for col_idx, k in enumerate(col_mapping.keys(), start=1):
        # 너비 대략 조정
        max_len = max((len(str(ws.cell(r, col_idx).value or "")) for r in range(1, ws.max_row+1)), default=12)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, int(max_len * 0.95)), 60)
        # 날짜 포맷
        if k == sort_by:
            for r in range(2, ws.max_row+1):
                c = ws.cell(r, col_idx)
                if isinstance(c.value, (datetime, date)):
                    c.number_format = "yyyy-mm-dd"
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=False, vertical="top")

    wb.save(path)
    return str(path)
# ================================ '엑셀에 데이터를 덧붙이는 함수' 모음 끝 ================================

# 엑셀 파일 -> PDF파일로 변환 (Windows + 로컬 Excel 필요)
def excel_to_pdf(
    filepath,
    output_path,
    source_sheet="전체",                 # 원본 시트명 (정확히)
    columns_order=None,                  # 출력 컬럼 순서 (기본값 아래 참조)
    small_headers=None,                  # 폭을 작게 고정할 헤더들
    big_headers=("문의내용", "답변"),     # 뒤쪽 반반으로 나눌 큰 컬럼들(줄바꿈 ON)
    small_col_widths=None,               # 작은 컬럼 폭(문자 단위) dict
    big_colwidth_ch=80,
    orientation="landscape",             # 'landscape' 또는 'portrait'
    repeat_header=True,
):
    """
    원본의 필요한 컬럼만(메타 + 문의내용/답변) 새 워크북으로 복사해
    작은 컬럼은 고정폭, '문의내용/답변'은 같은 폭으로 반반 배치하고 줄바꿈/행높이 자동맞춤 후 PDF 출력
    """
    # 기본 컬럼 순서(원하는 대로 바꿔도 됨)
    if columns_order is None:
        columns_order = ["쇼핑몰", "유형", "문의일", "답변여부", "작성자", "문의내용", "답변"]

    # 작은 컬럼 기본 세트
    if small_headers is None:
        small_headers = ["쇼핑몰", "유형", "문의일", "답변여부", "작성자"]

    # 작은 컬럼 폭 기본값(문자 단위)
    if small_col_widths is None:
        small_col_widths = {
            "쇼핑몰": 12,
            "유형": 10,
            "문의일": 12,
            "답변여부": 9,
            "작성자": 10,
        }

    # Excel 상수 (숫자값으로: makepy 캐시 의존 제거)
    xlTypePDF = 0
    xlQualityStandard = 0
    xlPortrait = 1
    xlLandscape = 2
    xlPaperA4 = 9
    xlContinuous = 1
    xlThin = 2
    # Borders index
    xlEdgeLeft, xlEdgeTop, xlEdgeBottom, xlEdgeRight = 7, 8, 9, 10
    xlInsideVertical, xlInsideHorizontal = 11, 12

    filepath = os.path.abspath(str(filepath))
    output_path = os.path.abspath(str(output_path))
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    pythoncom.CoInitialize()
    excel = None
    src_wb = None
    tmp_wb = None
    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False

        # 1) 원본 열 위치 매핑
        src_wb = excel.Workbooks.Open(filepath, ReadOnly=True)
        try:
            src_ws = src_wb.Sheets(source_sheet)
        except Exception:
            # 시트명이 다르면 첫 시트로 폴백
            src_ws = src_wb.Worksheets(1)

        used = src_ws.UsedRange
        first_col = used.Column
        last_col  = used.Column + used.Columns.Count - 1
        last_row  = used.Row + used.Rows.Count - 1

        header_index = {}
        for c in range(first_col, last_col + 1):
            h = src_ws.Cells(1, c).Value
            if h is not None:
                header_index[str(h).strip()] = c

        # 존재하는 컬럼만 필터링(순서는 columns_order 따름)
        selected_headers = [h for h in columns_order if h in header_index]

        if not selected_headers:
            raise RuntimeError("선택할 헤더를 찾을 수 없습니다. columns_order와 실제 헤더명을 확인하세요.")

        # 2) 임시 워크북 생성 & 데이터 복사
        tmp_wb = excel.Workbooks.Add()
        tws = tmp_wb.ActiveSheet
        tws.Name = "Print"

        # 헤더 행
        for j, h in enumerate(selected_headers, start=1):
            tws.Cells(1, j).Value = h

        # 각 컬럼을 배열로 복사(빠름)
        for j, h in enumerate(selected_headers, start=1):
            src_c = header_index[h]
            rng_src = src_ws.Range(src_ws.Cells(2, src_c), src_ws.Cells(last_row, src_c))  # 데이터만
            vals = rng_src.Value  # tuple-of-tuples or None (단일 칼럼이면 튜플)
            if vals is None:
                continue
            rng_dst = tws.Range(tws.Cells(2, j), tws.Cells(last_row, j))
            rng_dst.Value = vals

        # 3) 폭/줄바꿈/행높이/테두리 서식
        total_cols = len(selected_headers)

        # (A) 작은 컬럼: 줄바꿈 OFF + AutoFit + ShrinkToFit(보조)
        small_idxs = [j for j, h in enumerate(selected_headers, start=1) if h in small_headers]
        for j in small_idxs:
            col = tws.Columns(j)
            col.WrapText = False
            col.ShrinkToFit = True    # 아주 길 때 미세 축소
            try:
                col.AutoFit()         # 내용에 맞춰 폭 자동
            except Exception:
                pass

        # (A-1) 날짜 열(문의일) 형식 고정 + 다시 AutoFit → #### 방지
        for j, h in enumerate(selected_headers, start=1):
            if h == "문의일":  # 필요하면 다른 이름 추가
                col = tws.Columns(j)
                try:
                    col.NumberFormat = "yyyy-mm-dd"
                except Exception:
                    pass
                try:
                    col.AutoFit()
                except Exception:
                    pass

        # (B) 큰 컬럼(문의내용/답변): 줄바꿈 ON, 남은 폭을 반씩 정확히 배분
        big_idxs = [j for j, h in enumerate(selected_headers, start=1) if h in big_headers]
        # 우선 줄바꿈 켜고 임시 폭 지정(계산 전에 너무 작지 않게)
        for j in big_idxs:
            col = tws.Columns(j)
            col.WrapText = True
            col.ShrinkToFit = False
            col.ColumnWidth = 20

        # 페이지 설정(여백/방향)은 아래에서 다시 하지만, 폭 계산을 위해 미리 참조
        ps = tws.PageSetup
        is_land = orientation.lower().startswith("land")

        # (B-1) 페이지 가로 인쇄가능 폭(포인트) 계산
        page_width_pt = excel.Application.CentimetersToPoints(29.7 if is_land else 21.0)
        printable_pt  = page_width_pt - (ps.LeftMargin + ps.RightMargin)

        # (B-2) 현재 작은 컬럼 총폭(포인트) 합산
        def col_width_points(j):
            try:
                return float(tws.Columns(j).Width)  # points
            except Exception:
                return 0.0

        small_total_pt = sum(col_width_points(j) for j in small_idxs)

        # (B-3) 포인트↔문자 변환 비율 추정 (1문자당 포인트)
        sample_col = tws.Columns(1)
        try:
            points_per_char = sample_col.Width / max(sample_col.ColumnWidth, 1.0)
        except Exception:
            points_per_char = 7.2  # 대략값 (폴백)

        # (B-4) 남은 폭을 큰 컬럼 수로 균등 배분 → 문자인자로 환산해 ColumnWidth 지정
        remain_pt = max(printable_pt - small_total_pt, 100.0)  # 음수 보호 + 최소폭
        per_big_pt = remain_pt / max(len(big_idxs), 1)
        per_big_chars = max(per_big_pt / points_per_char, 15)  # 최소 15자

        for j in big_idxs:
            tws.Columns(j).ColumnWidth = per_big_chars

        # (C) 행 높이 자동맞춤(두 번 실행) → 줄바꿈 반영
        used2 = tws.UsedRange
        used2.Rows.AutoFit()
        used2.Rows.AutoFit()

        # (D) 얇은 테두리(전체 범위)
        rng_all = tws.Range(tws.Cells(1, 1), tws.Cells(last_row, total_cols))
        for idx in (xlEdgeLeft, xlEdgeTop, xlEdgeBottom, xlEdgeRight, xlInsideVertical, xlInsideHorizontal):
            b = rng_all.Borders(idx)
            b.LineStyle = xlContinuous
            b.Weight = xlThin
            b.ColorIndex = -4105  # Automatic

        # 4) 페이지 설정
        ps = tws.PageSetup
        ps.Orientation = xlLandscape if orientation.lower().startswith("land") else xlPortrait
        ps.PaperSize   = xlPaperA4
        ps.Zoom = False
        ps.FitToPagesWide = 1
        ps.FitToPagesTall = False
        ps.LeftMargin   = excel.Application.CentimetersToPoints(1.0)
        ps.RightMargin  = excel.Application.CentimetersToPoints(1.0)
        ps.TopMargin    = excel.Application.CentimetersToPoints(1.0)
        ps.BottomMargin = excel.Application.CentimetersToPoints(1.0)
        ps.CenterHorizontally = True
        ps.PrintGridlines = False
        if repeat_header:
            ps.PrintTitleRows = "$1:$1"

        # 인쇄 영역
        last_row2 = tws.UsedRange.Row + tws.UsedRange.Rows.Count - 1
        ps.PrintArea = tws.Range(tws.Cells(1, 1), tws.Cells(last_row2, total_cols)).Address

        # 5) PDF 출력
        tws.ExportAsFixedFormat(
            Type=xlTypePDF,
            Filename=output_path,
            Quality=xlQualityStandard,
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False,
        )
        return output_path
    finally:
        if tmp_wb:
            tmp_wb.Close(False)
        if src_wb:
            src_wb.Close(False)
        if excel:
            excel.ScreenUpdating = True
            excel.Quit()
        pythoncom.CoUninitialize()